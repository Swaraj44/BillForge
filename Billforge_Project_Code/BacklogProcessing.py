import openpyxl
import sys
import os

import time

import xlwings
import xlrd

# doc_path="demo"
# output_path="demo"
# email_path="demo"


import datasave as DS
import datashow as DSW
import pandas as pd
import tkinter as tk

import ConverterApi as tkconvert




#For Relative Resource Path
def resource_path(relative_path):
    """ Get absolute path to resource, works for dev and for PyInstaller """
    try:
        # PyInstaller creates a temp folder and stores path in _MEIPASS
        base_path = sys._MEIPASS2
    except Exception:
        base_path = os.path.abspath(".")

    return os.path.join(base_path, relative_path)






class Teacher:
    def __init__(self, name, designation, dept, question_paper_setter,Script_Examiner, examiners_class_tests, examiners_sessional_classes,
                 script_scrutinizer, tabulation, typing_and_drawing, central_viva_voce,
                 student_advising, seminar, thesis_progress_defense, final_grade_sheet_verification,
                 list_of_duty, sessional_course, theory_course):
        self.name = name
        self.designation = designation
        self.dept = dept
        self.question_paper_setter = int(question_paper_setter)
        self.Script_Examiner = int(Script_Examiner)
        
        self.examiners_class_tests = int(examiners_class_tests)
        self.examiners_sessional_classes = int(examiners_sessional_classes)
        self.script_scrutinizer = int(script_scrutinizer)
        self.tabulation = int(tabulation)
        self.typing_and_drawing = int(typing_and_drawing)
        self.central_viva_voce = int(central_viva_voce)
        self.student_advising = int(student_advising)
        self.seminar = int(seminar)
        self.thesis_progress_defense = int(thesis_progress_defense)
        self.final_grade_sheet_verification = int(final_grade_sheet_verification)
        self.list_of_duty = int(list_of_duty)
        self.sessional_course = sessional_course
        self.theory_course = theory_course

    def to_dict(self):
        return {
            'Name': self.name,
            'Designation': self.designation,
            'Dept': self.dept,
            'Question Paper Setter': self.question_paper_setter,
            'Script Examiner': self.Script_Examiner,
            'Examiners of Class Tests': self.examiners_class_tests,
            'Examiners of Sessional Classes': self.examiners_sessional_classes,
            'Script Scrutinizer': self.script_scrutinizer,
            'Tabulation': self.tabulation,
            'Typing and Drawing': self.typing_and_drawing,
            'Central Viva-Voce': self.central_viva_voce,
            'Student Advising': self.student_advising,
            'Seminar': self.seminar,
            'Thesis Progress Defense': self.thesis_progress_defense,
            'Final Grade Sheet Verification': self.final_grade_sheet_verification,
            'List of Duty': self.list_of_duty,
            'Sessional Course': self.sessional_course,
            'Theory Course': self.theory_course,
        }



teachers_array = []

from docx import Document
from docx.shared import Pt
from docx.enum.text import WD_UNDERLINE
#from googletrans import Translator
#module calling
import yearextractor as ys



TName = []


designation_map = {}
dept_map = {}
question_paper_setter_map = {}

Script_Examiner_map = {}

examiners_class_tests_map = {}


examiners_sessional_classes_map = {}

examiners_sessional_classes_bill={}

class_test_map = {}

script_scrutinizer_map = {}
tabulation_map = {}
typing_and_drawing_map = {}
central_viva_voce_map = {}
student_advising_map = {}
seminar_map = {}
thesis_progress_defense_map = {}
final_grade_sheet_verification_map = {}
list_of_duty_map = {}
sessional_course_map={}
sessional_course_Cradit_map={}
theory_course_map={}
Question_Paper_Moderation_Board ={}

# For Storing Multiple Sessional Course under o single teacher
teacher_sessional_course_collection_str ={}


teacher_sessional_courses_name={}


# For Storing Multiple Thoery ClassTest  Under a single Teacher

teacher_classtest_course_collection_str ={}

# For Storing Multiple Course Exampaper Evaluation under a single Teacher

teacher_exampaper_course_collection_str={}

Script_Examiner_course_collection_str={}


px = {}

bangla_teachers_name = {}

Course_Coordinator={}

Teachers_in_seminar_group={}


#####################################################################

Script_Examiner={}
Valid_Script_Examiner={}

student_per_subject={}

teacher_salary_id_map ={}

####################################################################


error_list = []







#####################################################
#Capturing Two Excel Values for out_of_scope use;
per_student_script_charge=0
minimum_script_charge=0
sessional_per_student_charge=0

payment_per_student_class_tests=0


def extract_data_from_excel():
    # Save any changes (if needed)
    workbook=openpyxl.load_workbook(resource_path('data\\excel-files\\templatebill.xlsx'))
    sheet = workbook.active
    
    global minimum_script_charge 
    global per_student_script_charge
    global payment_per_student_class_tests
    
    cell =sheet.cell(row=12, column=13)
    minimum_script_charge=int(cell.value)
    
    cell =sheet.cell(row=14, column=12)
    payment_per_student_class_tests=int(cell.value)
    
    
    cell =sheet.cell(row=12,column=12)
    per_student_script_charge=int(cell.value)
    
    cell =sheet.cell(row=17, column=12)
    global sessional_per_student_charge 
    sessional_per_student_charge=int(cell.value)

    # Optionally, delete the workbook object
    workbook.close()

extract_data_from_excel();




















bangla_teachers_designation ={
    'Dean' : 'ডীন',
    'Professor' : 'অধ্যাপক',
    'Associate Professor' : 'সহযোগী অধ্যাপক',
    'Assistant Professor' : 'সহকারী অধ্যাপক',
    'Lecturer' : 'প্রভাষক'
    
}

year_map = {
    '4th' : '৪র্থ',
    '3rd' : '৩য়',
    '2nd' : '২য়',
    '1st' : '১ম'
}

department_translation = {
    'DepartmentofCivilEngineering': 'সিভিল',
    'DepartmentofUrbanandRegionalPlanning': 'ইউআরপি',
    'DepartmentofBuilding Engineering & Construction Management ': 'বিইসিএম',
    'DepartmentofArchitecture': 'স্থাপত্য',
    'DepartmentofMathematics': 'গণিত',
    'DepartmentofChemistry': 'রসায়ন',
    'DepartmentofPhysics': 'পদার্থবিদ্যা',
    'DepartmentofHumanities': 'মানবিক',
    'DepartmentofElectricalandElectronicEngineering': 'ইইই',
    'DepartmentofComputerScienceandEngineering': 'সিএসই',
    'DepartmentofElectronicsandCommunicationEngineering': 'ইসিই',
    'DepartmentofBiomedicalEngineering': 'বিএমই',
    'DepartmentofMaterialsScienceandEngineering ': 'এমএসই',
    'DepartmentofMechanicalEngineering': 'মেকানিক্যাল',
    'DepartmentofIndustrialEngineeringandManagement ': 'আইইএম',
    'DepartmentofEnergyScienceandEngineering': 'ইএসই',
    'DepartmentofLeatherEngineering': 'লেদার',
    'DepartmentofTextileEngineering': 'টেক্সটাইল',
    'DepartmentofChemicalEngineering': 'কেমিক্যাল',
    'DepartmentofMechatronicsEngineering': 'মেকাট্রনিক্স'
}



department_translation_short = {
    'CE': 'সিভিল',
    'URP': 'ইউআরপি',
    'BECM': 'বিইসিএম',
    'ARCH': 'স্থাপত্য',
    'MATH': 'গণিত',
    'CHEM': 'রসায়ন',
    'PHY': 'পদার্থবিদ্যা',
    'HUM': 'মানবিক',
    'EEE': 'ইইই',
    'CSE': 'সিএসই',
    'ECE': 'ইসিই',
    'BME': 'বিএমই',
    'MSE': 'এমএসই',
    'ME': 'মেকানিক্যাল',
    'ESE': 'ইএসই',
    'LE': 'লেদার',
    'TE': 'টেক্সটাইল',
    'ChE': 'কেমিক্যাল',
    'MTE': 'মেকাট্রনিক্স',
    'IEM': 'আইইএম'
}


def is_underlined(run):
    """
    Check if a run contains underlined text.
    """
    return run.font.underline != WD_UNDERLINE.NONE

def find_table_heading(table):
    """
    Find the heading of a table by looking at the preceding paragraphs.
    """
    for paragraph in table._element.getprevious():
        if paragraph.tag.endswith("p"):
            if any(is_underlined(run) for run in paragraph.runs):
                return paragraph.text.strip()
    return None

#For Relative Resource Path
def resource_path(relative_path):
    """ Get absolute path to resource, works for dev and for PyInstaller """
    try:
        # PyInstaller creates a temp folder and stores path in _MEIPASS
        base_path = sys._MEIPASS2
    except Exception:
        base_path = os.path.abspath(".")

    return os.path.join(base_path, relative_path)




def translate_to_bengali(english_name):
    
    english_name = english_name.replace(" ", "")
    
    #translator = Translator()
    #translation = translator.translate(english_name, src='en', dest='bn')
    #return translation.text
    
    if english_name in department_translation_short:
        btext=department_translation_short[english_name]
    
    else:
       error_list.append("Error in Table 1 -->"+ english_name)
       error_list.append("Subjects name should be like-> CSE,EEE,PHY,CHEM,HUM,MATH,ChE,MTE,IEM")
       Error_view()
    
    
    return btext






def convert_to_bangla_year(year):
    # Bangla digits mapping
    bangla_digits = {'0': '০', '1': '১', '2': '২', '3': '৩', '4': '৪', '5': '৫', '6': '৬', '7': '৭', '8': '৮', '9': '৯'}

    # Convert the year to a string and replace each digit with its Bangla equivalent
    bangla_year = ''.join([bangla_digits[digit] for digit in str(year)])

    return bangla_year


def size(mapp):
    ck = 0
    for teacher in TName:
        cccc = mapp.get(teacher, 0)
        if cccc != 0:
        #   print(teacher)
        #   print(mapp[teacher])
          ck += 1
        

    return ck


lock=0;

def print_tables(doc):
    
    mailgenexcel = openpyxl.Workbook()
    sheet = mailgenexcel.active   
    cell=sheet.cell(row=1,column=1)
    cell.value ="Teachers' Name"
    cell=sheet.cell(row=1,column=4)
    cell.value ="Teachers' Email"
    flag=2
    
    
    global lock
    lock=lock+1
    
    
    for i, table in enumerate(doc.tables, start=1):
        
        if lock>1:
            break;
        # Extracting heading from the preceding paragraphs
        heading = find_table_heading(table)
        # Printing table heading
        #print(f"\nTable {i} Heading: {heading}")
        # Printing rows and values
        #print("-------------------Table :---------------------------"+str(i))

        if i == 1:
            #print("-------------------Table : 1---------------------------")
            r = 1
            for row in table.rows:
                data = "demo"
                data = table.rows[r].cells[1].text.strip()
                # print(data)

                data2 = "demo2"
                data2 = table.rows[r].cells[2].text.strip()
                # print(data2)

                if data2 == "Total =":
                    continue
                if data == "Total =":
                    continue
                result = data2.split(',')

                d2=result[0]
                d3=result[1]


                TName.append(data)
                
                #print("-----------------------------sw ============== ")
                #print(TName)
                #teacher_classtest_course_collection_str[data]=""
                
                
                
                designation_map[data] = d2
                dept_map[data] = d3
                Question_Paper_Moderation_Board[data] = 0

                r = r + 1;
                #global flag
                cell =sheet.cell(row=flag,column=1)
                cell.value=data
                flag=flag+1
                
                
                
                
                
                
                
                
        if i == 2:

            #print("-------------------Table : 2---------------------------")
            heading = find_table_heading(table)
            r = 1
            dm=0
            for row in table.rows:
                if dm==0:
                  dm=1
                  continue
                data = "demo"
                data = table.rows[r].cells[1].text.strip()
                #print(data)
                #print("---------------ck point1---------------------")

                data2 = "demo2"
                data2 = table.rows[r].cells[3].text.strip()
                #print("---------------ck point2---------------------")
                #print(data2)

                if data2 == "Total =":
                    continue
                if data == "Total =":
                    continue
                
                d2=0
                if data2=="Chairman" :
                    d2=1
                if data2=="Member" :
                    d2=2
                if data2=="Ext. Member" :
                    d2=2
                #data = translate_to_bengali[data]
                Question_Paper_Moderation_Board[data] = d2
                
                #print("------------data------------------------"+str(data))
                
                
                if data == '':
                    error_list.append("Warning: Empty Cell "+" in "+" table 2")
                else:           
                  if data not in TName:
                     error_list.append(data+" in "+" table 2")
                
                
                
                #print("Mumdu:"+str(r))

                #print(question_paper_setter_map[data])
                #print("------------------Mumdu  ST:----------------")
                #print(data)
                #print(Question_Paper_Moderation_Board[data])
                #print("------------------Mumdu  ED:----------------")


                r = r + 1;
            #print("-------------------ck-----------------")


        if i == 3:

            #print("-------------------Table : 3---------------------------")
            heading = find_table_heading(table)
            r = 1
            dm=0
            for row in table.rows:
                if dm==0:
                  dm=1
                  continue
                
                data_cr = "demo"
                data_cr = table.rows[r].cells[0].text.strip()
                
                data = "demo"
                data = table.rows[r].cells[1].text.strip()
                #print(data)
                
                
                data_r = "demo"
                data_r = table.rows[r].cells[2].text.strip()
                
                
                

                data2 = "demo2"
                data2 = table.rows[r].cells[3].text.strip()
                #print(data2)

                if data2 == "Total =":
                    continue
                if data == "Total =":
                    continue
                
                if data == '':
                    continue
                #data = translate_to_bengali[data]
                
                et = question_paper_setter_map.get(data, 0)
                question_paper_setter_map[data] = et+int(data2)
                #print("Printing ------------------------------------------------------------------Printing ",question_paper_setter_map[data],data)
                #print("Mumdu:"+str(r))
                
                
                ####
                student_per_subject[data_cr]=int(data2);
                
                
                
                
                
                if data_r == "Same as paper setter" or data_r == "Same as Paper Setter" or data_r == "same as paper setter":
                    et = Script_Examiner_map.get(data, 0)
                    Script_Examiner_map[data] = et+int(data2)
                    
                    et1 = Script_Examiner_course_collection_str.get(data, "")
                    if et1=="":
                            Script_Examiner_course_collection_str[data] =str(data_cr)
                    else:
                            Script_Examiner_course_collection_str[data] =et1+","+str(data_cr)
                            
                else:
                    et = Script_Examiner_map.get(data_r, 0)
                    Script_Examiner_map[data_r] = et+int(data2)
                    
                    et1 = Script_Examiner_course_collection_str.get(data_r, "")
                    if et1=="":
                            Script_Examiner_course_collection_str[data_r] =str(data_cr)
                    else:
                            Script_Examiner_course_collection_str[data_r] =et1+","+str(data_cr)
                    
                    
                    if data_r not in TName:
                     error_list.append(data_r+" in "+" table 3")
                    
                    
                    #Script_Examiner[data_r]=data2
                    #Valid_Script_Examiner[data_r]=1
                    #Valid_Script_Examiner[data]=-1
                    
                    
                
                    
                
                
                
                #print("--> Q Setter-Course :"+data)
                #print("--> Q Setter-Course stud numb :"+data2)
                
                
                #Script_Examiner_course_collection_str
                # Storing Multiple Quesion Paper and Script Examinar
                
                
                et1 = teacher_exampaper_course_collection_str.get(data, "")
                if et1=="":
                    teacher_exampaper_course_collection_str[data] =str(data2)
                else:
                    teacher_exampaper_course_collection_str[data] =et1+"+"+str(data2)
                

                #print(question_paper_setter_map[data])
                
                
                #print("------------------b****l:----------------")
                
                if data == '':
                    error_list.append("Warning: Empty Cell "+" in "+" table 3")
                else:           
                  if data not in TName:
                     error_list.append(data+" in "+" table 3")
                
                #print("--> Q DATA Name 3-1 :"+data)
                #print("--> Q DATA N 3-2 :"+data2)
                #print("--> Q DATA CN 3 :"+data3)
                
                #print("------------------Mumdu 111111:----------------")


                r = r + 1;
            #print("-------------------ck-----------------")

        if i == 4:

            #print("-------------------Table : 4---------------------------")
            heading = find_table_heading(table)
            r = 1

            dm=0
            for row in table.rows:
                if dm==0:
                  dm=1
                  continue
              
                data = "demo"
                data = table.rows[r].cells[1].text.strip()
                # print(data)

                data2 = "demo2"
                data2 = table.rows[r].cells[2].text.strip()

                data3 = "demo3"
                data3 = table.rows[r].cells[0].text.strip()
                
                
                data4 = "demo3"
                data4 = table.rows[r].cells[3].text.strip()

                

                if data2 == "Total =":
                    continue
                if data == "Total =":
                    continue
                if data3 == "Total =":
                    continue
                
                # print("--> Q DATA Name 1 :"+data)
                # print("--> Q DATA N 2 :"+data2)
                # print("--> Q DATA CN 3 :"+data3)
                
                
                #data44=1.5
                #data44=float(data4)
                #class_test_map[data]+=class_test_map[data]+data44;
                
                
                try:
                  data44 = float(data4)
                except ValueError:
                  data44 = 1.5
                  #print("Error: Could not convert to float. The string is empty or not a valid numerical value.")

                
                
                
                
                #print("DTAAAAAAAAA REMRK----->"+str(data44))
                
                
                et = class_test_map.get(data, 0)
                class_test_map[data] =et+data44
                
                
                
                
                
                et1 = teacher_classtest_course_collection_str.get(data, "")
                
                
                if et1=="":
                    teacher_classtest_course_collection_str[data] =str(data3)
                else:
                    teacher_classtest_course_collection_str[data] =et1+","+str(data3)
                
                
                if data not in theory_course_map:
                   theory_course_map[data] = []


                theory_course_map[data].append(data3)
                
                
                et = examiners_class_tests_map.get(data, 0)
                

                examiners_class_tests_map[data] =et+int(data2)
                #theory_course_map[data] = data3
                
                
                
                
                
                
                
                
                if data == '':
                    error_list.append("Warning: Empty Cell "+" in "+" table 4")
                else:           
                  if data not in TName:
                     error_list.append(data+" in "+" table 4")

                r = r + 1;

                #print(f"{data} is {data2}")





        
        if i == 5:
            print("-------------------Table : 5---------------------------")
            heading = find_table_heading(table)
            r = 1

            dm=0
            for row in table.rows:
                if dm==0:
                  dm=1
                  continue


                #print("------------------------------ggg----------------------------")
                data = "demo"
                data = table.rows[r].cells[1].text.strip()
                print(data)

                data2 = "demo2"
                data2 = table.rows[r].cells[2].text.strip()
                print(data2)
                data3 = 0.00
                data3 = float(table.rows[r].cells[3].text.strip())
                print(data3)


                # print(data2)
                data4 = "demo3"
                data4 = table.rows[r].cells[0].text.strip()
                

                if data2 == "Total =":
                    continue
                if data == "Total =":
                    continue
                #print("-----------------cccccc------------------------")
                
                
                et = examiners_sessional_classes_map.get(data, 0)
                examiners_sessional_classes_map[data] =et+int(data2)
                
                
                prev=examiners_sessional_classes_bill.get(data,0);
                st_num=int(data2);
                lab_cradit=data3;
                global sessional_per_student_charge;
                
                if data in examiners_sessional_classes_bill:
                  if lab_cradit==.75:
                    examiners_sessional_classes_bill[data]+=sessional_per_student_charge*st_num*.5;
                  elif lab_cradit==1.5:
                    examiners_sessional_classes_bill[data]+=sessional_per_student_charge*st_num*1;
                  elif lab_cradit==3:
                    examiners_sessional_classes_bill[data]+=sessional_per_student_charge*st_num*2;
                   
                else:
                    if lab_cradit==.75:
                      examiners_sessional_classes_bill[data]=sessional_per_student_charge*st_num*.5;
                    elif lab_cradit==1.5:
                      examiners_sessional_classes_bill[data]=sessional_per_student_charge*st_num*1;
                    elif lab_cradit==3:
                      examiners_sessional_classes_bill[data]=sessional_per_student_charge*st_num*2;
                
                
                
                
                #examiners_sessional_classes_map[data] = data2
                
                sessional_course_map[data]=data4
                
                
                
                
                sessional_course_Cradit_map[data4]=data3
                
                
                
                
                
                
                
                
            #     print("--> Q DATA Name 1  :"+data)
            #     print("--> Q DATA Number    :"+data2)
            #     print("--> Q DATA cradit     :"+str(data3))
            #     print("--> Q DATA CourseN     :"+str(data4))
            #    # print("Data 1 is ")
                #print(data)
                
                # #print("Data 2 is ")
                # # print(data2)
                # # print("Data 4 is ")
                # # print(data4)
                #print("Sessional Course Map ")
                #print(sessional_course_map[data])
                # print("Examiners Sessional Classes is ")
                # print(sessional_course_Cradit_map[data4])
                
                
                
                # For Stroing Multiple Sessional Courses
                
                et1 = teacher_sessional_course_collection_str.get(data, "")
                
                
                if et1=="":
                    teacher_sessional_course_collection_str[data] =str(data2)
                else:
                    teacher_sessional_course_collection_str[data] =et1+"+"+str(data2)
                    
                    
                    
                    
                
                et1 = teacher_sessional_courses_name.get(data, "")
                
                
                if et1=="":
                    teacher_sessional_courses_name[data] =str(data4)
                else:
                    teacher_sessional_courses_name[data] =et1+","+str(data4)
                
                
                
                
                
                if data == '':
                    error_list.append("Warning: Empty Cell "+" in "+" table 5")
                else:           
                  if data not in TName:
                     error_list.append(data+" in "+" table 5")

                r = r + 1;











        if i == 6:
            #print("-------------------Table : 5---------------------------")
            heading = find_table_heading(table)
            r = 1


            dm=0
            for row in table.rows:
                if dm==0:
                  dm=1
                  continue



                data = "demo"
                data = table.rows[r].cells[0].text.strip()
                # print("Printing Here")
                # print(data)

                data2 = "demo2"
                data2 = table.rows[r].cells[1].text.strip()
                # print("Printing Here V2")
                # print(data2)

                if data2 == "Total =":
                    continue
                if data == "Total =":
                    continue

                result = data2.split('=')

                d2=int(result[1])
                #print("Printing D2",d2)


                script_scrutinizer_map[data] = d2
                
                
                
                if data == '':
                    error_list.append("Warning: Empty Cell "+" in "+" table 6")
                else:           
                  if data not in TName:
                     error_list.append(data+" in "+" table 6")

                r = r + 1


        if i == 7:
            #print("-------------------Table : 6---------------------------")
            heading = find_table_heading(table)
            r = 1


            dm=0
            for row in table.rows:
                if dm==0:
                  dm=1
                  continue


                data = "demo"
                data = table.rows[r].cells[1].text.strip()
                # print(data)

                data2 = "demo2"
                data2 = table.rows[r].cells[2].text.strip()
                # print(data2)

                if data2 == "Total =":
                    continue
                if data == "Total =":
                    continue
                tabulation_map[data] = data2
                
                if data == '':
                    error_list.append("Warning: Empty Cell "+" in "+" table 7")
                else:           
                  if data not in TName:
                     error_list.append(data+" in "+" table 7")

                r = r + 1;


        if i == 8:
            heading = find_table_heading(table)
            r = 1


            dm=0
            for row in table.rows:
                if dm==0:
                  dm=1
                  continue



                data = "demo"
                data = table.rows[r].cells[0].text.strip()
                # print(data)

                data2 = "demo2"
                data2 = table.rows[r].cells[1].text.strip()
                # print(data2)

                if data2 == "Total =":
                    continue
                if data == "Total =":
                    continue
                typing_and_drawing_map[data] = data2
                
                
                
                if data == '':
                    error_list.append("Warning: Empty Cell "+" in "+" table 8")
                else:           
                  if data not in TName:
                     error_list.append(data+" in "+" table 8")

                r = r + 1;




        if i == 9:
            heading = find_table_heading(table)
            r = 1



            dm=0
            for row in table.rows:
                if dm==0:
                  dm=1
                  continue



                data = "demo"
                data = table.rows[r].cells[1].text.strip()
                # print(data)

                data2 = "demo2"
                data2 = table.rows[r].cells[2].text.strip()
                # print(data2)

                if data2 == "Total =":
                    continue
                if data == "Total =":
                    continue
                central_viva_voce_map[data] = data2
                
                if data == '':
                    error_list.append("Warning: Empty Cell "+" in "+" table 9")
                else:           
                  if data not in TName:
                     error_list.append(data+" in "+" table 9")

                r = r + 1;




        if i == 10:
            heading = find_table_heading(table)
            r = 1



            dm=0
            for row in table.rows:
                if dm==0:
                  dm=1
                  continue



                data = "demo"
                data = table.rows[r].cells[1].text.strip()
                # print(data)

                data2 = "demo2"
                data2 = table.rows[r].cells[2].text.strip()
                # print(data2)

                if data2 == "Total =":
                    continue
                if data == "Total =":
                    continue
                student_advising_map[data] = data2
                
                if data == '':
                    error_list.append("Warning: Empty Cell "+" in "+" table 10")
                else:           
                  if data not in TName:
                     error_list.append(data+" in "+" table 10")

                r = r + 1;


        if i == 11:
            heading = find_table_heading(table)
            r = 1



            dm=0
            for row in table.rows:
                if dm==0:
                  dm=1
                  continue

                data = "demo"
                data = table.rows[r].cells[1].text.strip()
                # print(data)

                data2 = "demo2"
                data2 = table.rows[r].cells[2].text.strip()
                # print(data2)
                
                number_of_teachers = "demo3"
                number_of_teachers = table.rows[r].cells[3].text.strip()
                
                
                

                if data2 == "Total =":
                    continue
                if data == "Total =":
                    continue
                seminar_map[data] = data2
                Teachers_in_seminar_group[data]=number_of_teachers;
                
                print("name-->",data,"---->",Teachers_in_seminar_group[data]);
                
                
                if data == '':
                    error_list.append("Warning: Empty Cell "+" in "+" table 11")
                else:           
                  if data not in TName:
                     error_list.append(data+" in "+" table 11")

                r = r + 1;







        if i == 12:
            heading = find_table_heading(table)
            r = 1


            dm=0
            for row in table.rows:
                if dm==0:
                  dm=1
                  continue



                data = "demo"
                data = table.rows[r].cells[1].text.strip()
                # print(data)

                data2 = "demo2"
                data2 = table.rows[r].cells[2].text.strip()
                # print(data2)
                
                
                

                if data2 == "Total =":
                    continue
                if data == "Total =":
                    continue
                thesis_progress_defense_map[data] = data2
                
                
                if data == '':
                    error_list.append("Warning: Empty Cell "+" in "+" table 12")
                else:           
                  if data not in TName:
                     error_list.append(data+" in "+" table 12")

                r = r + 1;












        if i == 13:
            heading = find_table_heading(table)
            r = 1


            dm=0
            for row in table.rows:
                if dm==0:
                  dm=1
                  continue



                data = "demo"
                data = table.rows[r].cells[1].text.strip()
                # print(data)

                data2 = "demo2"
                data2 = table.rows[r].cells[2].text.strip()
                # print(data2)

                if data2 == "Total =":
                    continue
                if data == "Total =":
                    continue

                #result = data2.split('(')
                #d2=int(result[0])


                final_grade_sheet_verification_map[data] = data2
                # print("Final Grade Sheet Verification: ")
                # print(data)
                # print(data2)
                # print(final_grade_sheet_verification_map[data])
                # print("...............")
                
                
                if data == '':
                    error_list.append("Warning: Empty Cell "+" in "+" table 13")
                else:           
                  if data not in TName:
                     error_list.append(data+" in "+" table 13")

                r = r + 1;
                
        if i == 14:
            heading = find_table_heading(table)
            r = 1


            dm=0
            for row in table.rows:
                if dm==0:
                  dm=1
                  continue



                data = "demo"
                data = table.rows[r].cells[1].text.strip()
                # print(data)

                data2 = "demo2"
                data2 = table.rows[r].cells[2].text.strip()
                # print(data2)

                if data2 == "Total =":
                    continue
                if data == "Total =":
                    continue
                
                Course_Coordinator[data] = 1
                
                if data == '':
                    error_list.append("Warning: Empty Cell "+" in "+" table 14")
                else:           
                  if data not in TName:
                     error_list.append(data+" in "+" table 14")
                     
                r = r + 1;



        if i == 15:
            heading = find_table_heading(table)
            r = 1


            dm=0
            for row in table.rows:
                if dm==0:
                  dm=1
                  continue



                data = "demo"
                data = table.rows[r].cells[1].text.strip()
                # print(data)

                data2 = "demo2"
                data2 = table.rows[r].cells[2].text.strip()
                # print(data2)

                if data2 == "Total =":
                    continue
                if data == "Total =":
                    continue
                list_of_duty_map[data] = data2
                
                if data == '':
                    error_list.append("Warning: Empty Cell "+" in "+" table 15")
                else:           
                  if data not in TName:
                     error_list.append(data+" in "+" table 15")

                r = r + 1;
    global output_path            
    directory_path = output_path
    
    #file_name = '11111names_and_emails.xlsx'
    #full_path = f'{directory_path}/{file_name}'
    full_path =email_path

    mailgenexcel.save(full_path)
    
    


def create_teacher(i):
    name = TName[i]
    designation = designation_map[name]
    dept = dept_map[name]


    question_paper_setter = question_paper_setter_map.get(name, 0)
    
    ####
    Script_Examiner = Script_Examiner_map.get(name, 0)
    
    
    examiners_class_tests = examiners_class_tests_map.get(name, 0)
    examiners_sessional_classes = examiners_sessional_classes_map.get(name, 0)
    script_scrutinizer = script_scrutinizer_map.get(name, 0)
    tabulation = tabulation_map.get(name, 0)
    typing_and_drawing = typing_and_drawing_map.get(name, 0)
    central_viva_voce = central_viva_voce_map.get(name, 0)
    student_advising = student_advising_map.get(name, 0)
    seminar = seminar_map.get(name, 0)
    thesis_progress_defense = thesis_progress_defense_map.get(name, 0)
    final_grade_sheet_verification = final_grade_sheet_verification_map.get(name, 0)
    list_of_duty = list_of_duty_map.get(name, 0)
    sessional_course = sessional_course_map.get(name, "Null")
    
    list1=theory_course_map.get(name, "Null")
    
    s=""
    for i in list1:
        s+=i+" "
        
    
    theory_course = s
    


    # Creating and returning the Teacher instance
    return Teacher(name, designation, dept, question_paper_setter,Script_Examiner, examiners_class_tests,
                   examiners_sessional_classes, script_scrutinizer, tabulation, typing_and_drawing,
                   central_viva_voce, student_advising, seminar, thesis_progress_defense,
                   final_grade_sheet_verification, list_of_duty, sessional_course, theory_course)



def excel_add(cteacher):
    workbook = openpyxl.load_workbook(resource_path('data\\excel-files\\templatebill_backlog.xlsx'))
    name_string = "নাম: "
    podobi_string = "পদবী: "
    bivag_string = "বিভাগ/শাখা: "
    # Select the active sheet

    sheet = workbook.active
    
    
    # Writing Name to the excel Sheet Demo( Name: Dr Sheikh Imran Hossain)
    merged_cell = sheet[ 'A3:C3' ]
    new_value = cteacher.name
    #/////////////////////////////////////////////////////////////////////////Turjo///////////////////////////////////////////////////
    st1 = "Null"
    if cteacher.name in bangla_teachers_name:
      st1 = bangla_teachers_name[cteacher.name]

    
    if(st1!="Null"):
         new_value = bangla_teachers_name[cteacher.name]
    #print("New Value is : ", new_value)
         merged_cell[0][0].value = name_string + new_value
    else:
        temp_str="Bangla Teacher's Name Mismatching! "+cteacher.name
        error_list.append(temp_str)
        Error_view()
    
        
        
    
    new_value = bangla_teachers_name[cteacher.name]
    #print("New Value is : ", new_value)
    merged_cell[0][0].value = name_string + new_value
    
    
    
    # Writing Designation to the excel sheet Demo(Name: Assistant Professor)
    merged_cell = sheet[ 'A4:C4' ]
    new_value = bangla_teachers_designation[cteacher.designation]
    #print("New Value is : ", new_value)
    merged_cell[0][0].value = podobi_string + new_value
    
    
    # Writing Department to the excel sheet Demo (Dept: CSE )
    
    merged_cell = sheet[ 'A5:B5' ]
    new_value = cteacher.dept
    dept_name_bengali = translate_to_bengali(new_value)
    merged_cell[0][0].value = bivag_string + dept_name_bengali
    
    # Question Paper Setter Info to excel
    
    
    
    if cteacher.name in teacher_salary_id_map:
            
        cell = sheet.cell(row=3, column=5)
        z=teacher_salary_id_map[cteacher.name]
        cell.value = z
    
   
    
    # Writing BSC ENG Name to the excel file
    
    
    #EEEEEEEEEEEEEEE Eikhane Change Korte Hobe EEEEEEEEEEEEEEEEEEEEEE
    for row in sheet.iter_rows(min_row=9, max_row=9, min_col=3, max_col=3):
        for cell in row:
            cell.value = "বি. এসসি. ইঞ্জি:"
            
    for row in sheet.iter_rows(min_row=10, max_row=30, min_col=3, max_col=3):
        for cell in row:
            cell.value = '"'
            
  
            
            
   
       
       
    # Question Niamon
        
    # cell = sheet.cell(row=11, column=7)
    # cell.value = 2  
    
    
    
    
    
    # Question Examinar
    #cell = sheet.cell(row=12, column=7)
    #cell.value = cteacher.question_paper_setter
    
    cell = sheet.cell(row=12, column=7)
    cell.value = cteacher.Script_Examiner
   
    
    
    
    
    #Clss Tests
    cell = sheet.cell(row=14, column=7)
    cell.value = cteacher.examiners_class_tests
    
    #Errornous Code:
    # if cteacher.name in teacher_classtest_course_collection_str:
    #     str1=teacher_classtest_course_collection_str[cteacher.name] + "="+str(cteacher.examiners_class_tests)
    #     cell.value = str1
    
    
    # Seminar 
    cell = sheet.cell(row=16, column=7)
    temp1=cteacher.seminar
    cell.value = cteacher.seminar

    
    
    #Total Teacher Attending Seminars
  
    if temp1 >0:
          cell = sheet.cell(row=16, column=8)
          v1 =Teachers_in_seminar_group[cteacher.name]
          cell.value=v1
    
    
    


    
    
    ########################################################################################################################
    
    '''
    cell = sheet.cell(row=12, column=5)
    if(cteacher.theory_course!="Null"):
     
     z=cteacher.theory_course
     cell.value = z
     
     '''
     
    if cteacher.name in Script_Examiner_course_collection_str:
            
            cell = sheet.cell(row=12, column=5)
            z=Script_Examiner_course_collection_str[cteacher.name]
            cell.value = z
            
            
            ##########################################################
            
            total=0
            total_std="("
            cnt=0
            
            sub_arr1=Script_Examiner_course_collection_str[cteacher.name]
            #Reading Cell Value From How many tk per khata one would get
            cell = sheet.cell(row=12, column =12)
            tb1=cell.value
            sub_arr = sub_arr1.split(',')
            
            #print("-------------------------------------------->"+str(sub_arr))
            for xx in sub_arr:
                 if xx ==',':
                     continue
                 dd=student_per_subject[xx]
                 if cnt==0:
                     #print("Jk---11------>>>>>"+total_std+"         --->"+xx)
                     cnt=cnt+1
                     total_std=total_std+str(dd)
                     
                 else:
                     #print("Jk---22------>>>>>"+total_std+"         --->"+xx)
                     total_std=total_std+"+"+str(dd)
                     
                 
                 
                 if(dd>8):
                     total=total+int(tb1)*dd
                     
                 else:
                     cell = sheet.cell(row=12, column= 13)#reading 750
                     tb2 =cell.value
                     total=total+ tb2
                     
                 
            total_std=total_std+")"    
                     
           
                     
            
            cell =sheet.cell(row=12,column=10)
            cell.value = total;        
            
            
            cell =sheet.cell(row=12, column =7)
            #total_std="Turjo"
            cell.value =str(total_std)
                     
            
            
            
            
            
            
     
     ##########################################################################################################################
    
    #if(cteacher.theory_course!="Null"):
    st1="Null"
    #st1 = teacher_classtest_course_collection_str.get(cteacher.name, "Null")
    #st1 = teacher_classtest_course_collection_str[cteacher.name]
    
    st1 = "Null"
    if cteacher.name in teacher_classtest_course_collection_str:
      st1 = teacher_classtest_course_collection_str[cteacher.name]


    
    
    
    if(st1!="Null"):
       
        
        #teacher_classtest_course_collection_str
        cell = sheet.cell(row=9, column=7)
        
        
        
        #s1=cteacher.theory_course
        s1=teacher_classtest_course_collection_str[cteacher.name]
        result = s1.split(',')

        d2=int(len(result))
        
                                                                                      # Unmerge the cell before setting the value
        cell.value =  d2
        
        
        
        cell = sheet.cell(row=9,column=5)
        cell.value = teacher_classtest_course_collection_str[cteacher.name]
        
        
        
        
        #########################################################
        #cell = sheet.cell(row=12, column=5)
        #z=cteacher.theory_course
        #cell.value = z
        

            
        
        
        
        
        cell = sheet.cell(row=14, column=5)
        #z=cteacher.theory_course
        z=teacher_classtest_course_collection_str[cteacher.name]
        cell.value = z
        
        
    
        
    
    #Class Test Subject
    ## Handling Null Values
    
    '''
    cell = sheet.cell(row=14, column=5)
    if(cteacher.theory_course!="Null"):
     
     z=cteacher.theory_course
     cell.value = z
    '''
    #Class Test Number
    
    if cteacher.name in class_test_map:
    
       cell = sheet.cell(row=14, column=8)
       cell.value = class_test_map[cteacher.name]
    
    
    

    #sessional_per_student_charge
    if cteacher.name in examiners_sessional_classes_bill:
        cell = sheet.cell(row=17, column=10)
        cell.value = examiners_sessional_classes_bill[cteacher.name]
        
    
    y = cteacher.sessional_course
    cell = sheet.cell(row=17, column=8)
    if(y != "Null"):
        x = sessional_course_Cradit_map[y]    
        
        cell.value =  float(x)
        #print(cell.value)
    
    
    
        
    if cteacher.name in teacher_sessional_course_collection_str :
        
        cell = sheet.cell(row=17, column=7)
        z=teacher_sessional_course_collection_str[cteacher.name]
        cell.value=z
        
        
    if cteacher.name in teacher_sessional_courses_name :
        
        cell = sheet.cell(row=17, column=5)
        z=teacher_sessional_courses_name[cteacher.name]
        cell.value=z
        
        
    
    
        
  
    
     
    #Central Viva (Total Teacher + Total Student)
    #Modifided For V_3.3
    cell = sheet.cell(row=18, column=7)
    cell.value = cteacher.central_viva_voce
    #tp1=size(central_viva_voce_map)
    #cell.value = tp1
    temp1 = cteacher.central_viva_voce
    
    
    cell = sheet.cell(row=18, column=8)
    ck=  size(central_viva_voce_map)
    if ( temp1 !=0):
        cell.value = ck
    else:
        cell.value = 0
   
    
    
    
    #Thesis Progress Defense
    #Modified For V_3.3
    cell = sheet.cell(row=20, column=7)
    cell.value = cteacher.thesis_progress_defense
    temp1= cteacher.thesis_progress_defense
    
    #Here temp1 variable is used for checking whether the left side value is zero 
    #or not if zero then this cell value will be zero
    cell = sheet.cell(row=20, column=8)
    ck = size(thesis_progress_defense_map)
    if (temp1 != 0):
        cell.value = ck
    else:
        cell.value = 0
    
    
    #Modified For V_3.3
    cell = sheet.cell(row=24, column=7)
    cell.value = cteacher.tabulation
    #Tabulation and Tabulation Verification Both are same
    #That's why we are stroing the same value to the both columns
    cell =sheet.cell(row=23, column =7)
    cell.value = cteacher.tabulation
    
    
    #Scrutinizer
    cell = sheet.cell(row=25, column=7)
    cell.value = cteacher.script_scrutinizer
    
    
    # List of Duty
    cell = sheet.cell(row=26, column=7)
    cell.value = cteacher.list_of_duty
    
    
    #Typing & Drawing
    
    cell = sheet.cell(row=27, column=7)
    cell.value = cteacher.typing_and_drawing
    
    
    
    cell = sheet.cell(row=27, column=8)
    cell.value= cteacher.typing_and_drawing
    
    
    #Final Grade Sheet Verification
    
    cell = sheet.cell(row=28, column=7)
    cell.value= cteacher.final_grade_sheet_verification
    
    
    
    cell = sheet.cell(row=28, column=8)
    cell.value= cteacher.final_grade_sheet_verification
    
    #Student Advising
    
    cell = sheet.cell(row=29, column=7)
    cell.value = cteacher.student_advising
    
    
    
    
    #######################################################################################-----------> Sw
    st1 = 0
    if cteacher.name in Course_Coordinator:
      st1 = Course_Coordinator[cteacher.name]

    
    if(st1!=0):
        cell = sheet.cell(row=31, column=7)
        cell.value = st1
    
    # Taka Amount
  
    #################Time is USED HERE###################
    #time.sleep(1)
    
    
    cell = sheet['J32']
    
    taka =cell.value
    #print(cell.value)
    
   
    
    
    
    
    #merged_cell = sheet[ 'A32:E32' ]
    
    #merged_cell[0][0].value = taka
    
    
    
    # Year & Semester Setter
    # Year & Name Setter From docx from excel
    #str2=ys.yearsemextractor()
    str2=ys.yearsemextractor(doc_path)
    result = str2.split('-')
    d2=result[1]
    Examyear=result[2]

    #print(d2)
    # Examyear = translate_to_bengali(Examyear) 
    # print("ExamYear is: ")
    # print(Examyear)# Google Translator API
    
    Examyear = convert_to_bangla_year(Examyear)
    
    merged_cell = sheet[ 'F3:I3' ]
    merged_cell[0][0].value = "ব্যাকলগ পরীক্ষা "+str(Examyear)


    #result2 = d2.split(' ')

    #year=result2[1]
    #term=result2[3]
    
    # Putting the Bangla Version of Year (2022)
    cell = sheet.cell(row=4, column=7)
    #cell.value = year_map[year]
    cell.value = "সকল"
    
    
    # Printing the value of year in the excel sheet
    for row in sheet.iter_rows(min_row=9, max_row=9, min_col=4, max_col=4):
        for cell in row:
            cell.value = Examyear
            
    for row in sheet.iter_rows(min_row=10, max_row=30, min_col=4, max_col=4):
        for cell in row:
            cell.value = '"'
    
    # Putting the Bangla Version of Term(1st)
    cell = sheet.cell(row=4, column=9)
   
    #cell.value = year_map[term]
    cell.value = "সকল"
    
    merged_cell = sheet['F5:I5']
    dept_name = ys.nameextractor(doc_path)
    merged_cell[0][0].value = merged_cell[0][0].value + department_translation[dept_name] 
    
    
    
    # Moderation Board Calculation
    #Here we are assuming that in every semester there are 5 courses and each moderation board cons
    #consists of 5 person
    if (Question_Paper_Moderation_Board[cteacher.name]== 1):
        cell =sheet.cell(row=10, column=7)
        cell.value=1
        cell =sheet.cell(row=11,column=7)
        #previously this value was 5
        cell.value=1
        ck =size(Question_Paper_Moderation_Board)
        
        # print("Moderation Board size is :......")
        # print(cteacher.name)
        # print(Question_Paper_Moderation_Board[cteacher.name])
        # print(ck)
        ck = int(ck)
        cell =sheet.cell(row=11,column=10)
        moderate_taka = float((10*3600)/(1.00*ck))
        if( moderate_taka >= 4500):
            cell.value = 4500
        elif(moderate_taka <=1500): #Added the below threashold value for moderate taka
            cell.value = 1500
        else:
            cell.value = moderate_taka
    elif (Question_Paper_Moderation_Board[cteacher.name]== 2):
        cell =sheet.cell(row=11,column=7)
        cell.value=1
        ck =size(Question_Paper_Moderation_Board)
        ck = int(ck)
        cell =sheet.cell(row=11,column=10)
        moderate_taka = float((10*3600)/(1.00*ck))
        if( moderate_taka >= 4500):
            cell.value = 4500
        elif(moderate_taka <=1500):
            cell.value =1500
        else:
            cell.value = moderate_taka
            
    else:
        dummy=10 #God Knows what does this variable do! maybe in future
        #this is maybe of some use
        
    
   
    
    
    
    

#################################################################
    
    # Saving the name to the respective teachers
    
    
    file_name = f"{cteacher.name}.xlsx"
    cpath=output_path+"/"+file_name
    #workbook.save(f'D:/Academic/System Final/Solution/{file_name}')
    workbook.save(cpath)
    
    ###############Timer Is USed here################
    time.sleep(0.001)
    excel_app = xlwings.App(visible=False)
    #excel_book = excel_app.books.open(f'D:/Academic/System Final/Solution/{file_name}')
    excel_book = excel_app.books.open(cpath)
    excel_book.save()
    excel_book.close()
    excel_app.quit()
    #loaded_workbook = openpyxl.load_workbook(f'D:/Academic/System Final/Solution/{file_name}', data_only=True)
    loaded_workbook = openpyxl.load_workbook(cpath, data_only=True)
    sheet = loaded_workbook.active
    
    loaded_workbook.calculate_before_save = True
    
    
    #print("Values ....")
    print(cteacher.name)
    #print(sheet['I32'].value)
    #print(type(sheet['I32'].value))
    
    
    takavalue= round(float(sheet['J32'].value),2)
    formatted_value = "{:.2f}".format(takavalue)
    ##########################################################################################################################################################
    
    #DS.savedata(cteacher.name,formatted_value)
    DS.savedata2(cteacher.name,formatted_value,"summarydb")
    
    
    
    takavalue= str(formatted_value)
    #print("Before Going TO Api: ")
    #print(takavalue)
    
    #taka =apc.cur2bangla(takavalue)
    
    taka =tkconvert.Number_to_Money(takavalue)
    
    #####################Timer is USED HERE *********************
    time.sleep(0.0001)
    #print("Value from api is: ...................")
    taka = "কথায়: "+taka
    #print(taka)
    
    loaded_workbook.save(cpath)
    
    
    workbook = openpyxl.load_workbook(cpath)
    sheet = workbook.active
    merged_cell = sheet[ 'A32:B32' ]
    merged_cell[0][0].value = taka
    
    workbook.save(cpath)
    
 ################################################################################################   







def fetchD(uname):
    ss=DSW.showdata(uname)
    print(ss)
    #return ss


    
    
    #print(cteacher.designation)
    
    
def summarygenerator():
    
     str2=ys.yearsemextractor(doc_path)
     
     summary = openpyxl.Workbook()
     sheet = summary.active
     
     merged_cell = sheet ['B2:F2']
     merged_cell[0][0].value=str2 
     
     merged_cell = sheet['B3:C3']
     merged_cell[0][0].value = "Teacher's Name"
     merged_cell = sheet['D3:F3']
     merged_cell[0][0].value = "Teacher's Bill" 
     sum1 = 4
     total_bill=0.0
    
     
     for teacher in teachers_array:

        print("Name: "+teacher.name)
        ss=DSW.showdata2(teacher.name,"summarydb") # Fetching  Data From Database Showing the teacher's total amount
        print("value: "+ss)
        cell = sheet.cell(row=sum1,column=2)
        cell.value = teacher.name
        cell = sheet.cell(row=sum1,column=4)
        cell.value = ss
        sum1=sum1+1
        total_bill = total_bill + float(ss)
        
     cell = sheet.cell(row=sum1,column=4)
     cell.value = total_bill
     cell = sheet.cell(row=sum1,column=2)
     cell.value = "Total Bill For Teachers: "
     save_name = output_path+"/"+str2+"Summary.xlsx"
     print("Summary Generated....................")
     
     summary.save(save_name)





def Error_view():
    # Create the Tkinter window
    root = tk.Tk()
    root.title("Errors")

    # Determine the maximum length of sentence in the error list
    max_sentence_length = max(len(sentence) for sentence in error_list)

    # Calculate a suitable width based on the maximum sentence length
    width = max_sentence_length + 10  # Adding some padding

    # Create a Listbox widget with a specified width
    listbox = tk.Listbox(root, height=len(error_list), width=width)
    listbox.pack(padx=10, pady=10)  # Add padding around the listbox

    # Add items to the Listbox
    for item in error_list:
        listbox.insert(tk.END, item)

    # Function to close the window
    def close_window():
        root.destroy()

    # Create a Close button
    close_button = tk.Button(root, text="Close", command=close_window)
    close_button.pack(pady=10)

    # Start the Tkinter event loop
    root.mainloop()





def main():
    global doc_path
    
    doc_path1 = doc_path

    try:
        document = Document(doc_path1)
        print_tables(document)
        
    except Exception as e:
        print(f"An error occurred: {e}")
        
        



    num_teachers = len(TName)
    
    ck = size(sessional_course_map)
    # print("Sessional size is: ")
    # print(ck)

# Creating instances and adding them to the array
    for i in range(num_teachers):
        teacher_instance = create_teacher(i)
        teachers_array.append(teacher_instance)






##############################################################################


# Accessing and printing information from the array of teachers
    for teacher in teachers_array:

        print(teacher.name)
        #print(TName)
        
        
        
        nameee=teacher.name
        #print(teacher.sessional_course)
        if nameee in TName:
           excel_add(teacher)
           
           #print(teacher.to_dict())
        
        
        
   
        
        
        
        #print(teacher.sessional_course)
        #excel_add(teacher)
        #print(teacher.to_dict())
        
        
    #summarygenerator()
    
    
 
    
    if len(error_list) != 0:
        print(error_list)
        Error_view()
        
    else:
        error_list.append("Bill Is Generated without Error")
        Error_view()
    
    
    

    
    
    
    print("Code Executed Successfully")                
                  
                  
      
      
def call_to_main(dpath,epath,opath):
    global doc_path
    global output_path
    global email_path
    
    doc_path=dpath
    output_path=opath
    email_path=epath
    
    #doc_path='special backlog.docx'
    #output_path='Test Gen'
    
    #email_path='C:/Users/DELL/Desktop/SYS Project/Updated/System Final/Turjo/11111names_and_emails.xlsx' 
    
    df = pd.read_excel(resource_path('data\\excel-files\\bangla_teachers_name.xlsx'))

    # Extract columns 1 and 2 (0-indexed) and iterate over rows
    for index, row in df.iloc[:, :3].iterrows():
    # Access data of columns 1 and 2
    
     name = row[0]
     b_name = row[1]
     salary_id = row[2]
     
     bangla_teachers_name[name]=b_name
     teacher_salary_id_map[name] = salary_id
     
     #print(bangla_teachers_name);

    main()
            
if __name__ == "__main__":
    
    #main()
    df = pd.read_excel(resource_path('data\\excel-files\\bangla_teachers_name.xlsx'))
    
    #bangla_teachers_name = {}
    

    # Extract columns 1 and 2 (0-indexed) and iterate over rows
    for index, row in df.iloc[:, :2].iterrows():
    # Access data of columns 1 and 2
     name = row[0]
     b_name = row[1]
     
     bangla_teachers_name[name]=b_name;
    #print(bangla_teachers_name);
     
    
    # Print data
     #print("Column 1:", name)
     #print("Column 2:", b_name)
     
    #print(bangla_teachers_name)
    
    call_to_main("aa","bb","cc")
    

    
    
    